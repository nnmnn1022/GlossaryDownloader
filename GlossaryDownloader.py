import os
import csv
import gspread
import pathlib
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import load_workbook
from os import walk
from setPath import run as setPath
from setPath import setUniqFileName

def run() :
    print("============================== [START] 시작 ==============================")
    # 세팅 불러오기
    setting = load_setting()

    # 미니 용어집 불러오기
    path = setPath(input("\n***** [Notice] 미니 용어집을 드래그하거나 경로(.csv)를 입력해 주세요.*****\n> "))
    mGlossary = loadMGlossary(path, setting)
    mData = mGlossary[0]
    eFile_col_info = setting[10]
    extra_eFile_col_info = setting[13:]
    target = mData[1].upper().strip()

    # 22.10.27 요청으로 인한 수정 - 구글시트 링크가 없으면 생략하고 진행
    # 구글시트와 비교 / 구글시트 링크가 없으면 생략하고 진행
    print("구글 용어집 확인 중..")
    if setting[4]:
        googleGlossaryPath = setting[4][0]
        gGlossary_col_info = setting[7]
        gData = loadGsheet(target, googleGlossaryPath)
        compared_data = gCompare(gData, mGlossary, gGlossary_col_info)
    else :
        print("구글 용어집 링크가 존재하지 않습니다.\n스텝을 생략합니다.")
        compared_data = mGlossary

    # 엑셀과 비교
    eData = loadEsheet(eFile_col_info, extra_eFile_col_info)
    compared_data = Ecompare(eData, compared_data)

    # 엑셀에서 단어가 들어간 내용 찾기
    compared_data = Efind(eData, compared_data)
    
    # 생성되는 파일 이름 관리
    dirName, changedName  = setUniqFileName(path, target)

    writeFile(compared_data, changedName, dirName)

    print("============================== [END] 완료 ==============================")
    os.system("pause")


def loadEsheet(eFile_col_info, extra_eFile_col_info) :
    data = []
    path = setPath(input("\n***** [Notice] 엑셀 파일 경로 또는 폴더의 경로를 입력해 주세요.*****\n> "))
    # path = '''D:\\!Project\\Umoo\\umoo\\Wemade\\1_Work\\0209_8테마선번역_Regular\\ㅗㅜㅑ'''

    if '.xlsx' in os.path.splitext(path)[1].lower() :
        filename = path.split('/')[-1]
        loadedFile = load_workbook(path, data_only=True)
        sheetNames = loadedFile.sheetnames
        # 엑셀 파일 소스열 설정
        source_col = eFile_col_info[0]
        # 엑셀 파일 타겟열 설정
        target_col = eFile_col_info[1]

        # 현재 엑셀 파일이름 안에 예외 파일들의 이름이 포함되어 있으면 소스/ 타겟열을 그에 맞게 바꿔주기
        if extra_eFile_col_info and extra_eFile_col_info != [[]]:
            for extra_info in extra_eFile_col_info:
                if extra_info[0] in filename:
                    source_col = extra_info[1]
                    target_col = extra_info[2]

        for sheetName in sheetNames:
            sheet = loadedFile[sheetName]

            for row in sheet.iter_rows(min_row=2):
                row_value = [str(row[int(source_col)].value), str(row[int(target_col)].value)]
                if None not in row_value :
                    data.append(row_value)

    else :
        # 폴더의 모든 파일 가져오기
        targetFilesAbsolutePaths = []
        for (dirPaths, dirNames, fileNames) in walk(path):
            targetFilesAbsolutePaths.extend([dirPaths + '\\' + fileName for fileName in fileNames])

        for fileAbsolutePath in targetFilesAbsolutePaths:
            # 엑셀 파일이 아니면 다음 내용들을 수행하지 않음
            if os.path.splitext(fileAbsolutePath)[1].lower() != ".xlsx": continue
            # 확장자를 포함한 파일이름
            filename = fileAbsolutePath.split('/')[-1]
            loadedFile = load_workbook(fileAbsolutePath, data_only=True)
            sheetNames = loadedFile.sheetnames
            # 조건에 따라 소스열, 타겟열 받아오기
            source_col = eFile_col_info[0]
            target_col = eFile_col_info[1]

            # 현재 엑셀 파일이름 안에 예외 파일들의 이름이 포함되어 있으면 소스/ 타겟열을 그에 맞게 바꿔주기
            if extra_eFile_col_info and extra_eFile_col_info != [[]]:
                for extra_info in extra_eFile_col_info:
                    if extra_info[0] in filename:
                        source_col = extra_info[1]
                        target_col = extra_info[2]

            for sheetName in sheetNames:
                sheet = loadedFile[sheetName]
                try:
                    for row in sheet.iter_rows(min_row=2):
                        row_value = [str(row[int(source_col)].value), str(row[int(target_col)].value)]
                        if None not in row_value : data.append(row_value)
                except IndexError:
                    print('[Error] 엑셀 파일의 한글, 타겟열 설정이 잘못되었거나,')
                    print('[Error] 설정보다 적은 열을 가지고 있는 엑셀 파일이 포함되었습니다.')
                    print('[Error] 프로그램을 종료합니다.')
                    os.system("pause")
                    exit()

    print("엑셀 파일 확인 완료")

    return(data)


def loadGsheet(target, googleGlossaryPath) :
    doc = connectGsheet(googleGlossaryPath)
    worksheets = doc.worksheets()
    sheetname_list = [worksheet.title for worksheet in worksheets]
    if target in sheetname_list :
        worksheet_index = sheetname_list.index(target)
    else :
        # map + lambda 보다 list comprehesion이 더 성능이 좋음
        sheetname_list = [sheetname.upper().replace('KO2','') for sheetname in sheetname_list]
        if target in sheetname_list :
            worksheet_index = sheetname_list.index(target)
        else : print('[Error] Target 언어 용어집을 찾을 수 없습니다.')
    
    worksheet = doc.get_worksheet(worksheet_index)
    gdata = worksheet.get_all_values()

    print("구글 용어집 확인 완료")
    return gdata


def gCompare(gGlossaryData, miniGlossaryData, gGlossary_col_info) :
    print("텀베이스 용어집과 미니 용어집 비교 중...")
    # 해당 언어 전체 구글 시트 내용
    gData = gGlossaryData
    # 미니 용어집의 한글 + 해당 언어열
    mData = miniGlossaryData
    gSelected_source_col = int(gGlossary_col_info[0])
    gSelected_target_col = int(gGlossary_col_info[1])
    data = [mData[0]]

    for mRow in mData[1:] :
        # 같은 한글이 있는지 확인
        for gRow in gData[1:] :
            row_data = [mRow[0]]
            if len(mRow) > 1 and mRow[-1] != '':
                row_data.extend([mRow[1], '√'])
                break

            elif (mRow[0].replace(' ', '') == gRow[gSelected_source_col].replace(' ', '')) and gRow[gSelected_target_col] != '' :
                row_data.extend([gRow[gSelected_target_col], '√'])
                break

        data.append(row_data)
    print("비교 완료")
    return data


def Ecompare(eGlossaryData, miniGlossaryData) :
    print("엑셀 파일과 미니 용어집 비교 중...")
    eData = eGlossaryData
    mData = miniGlossaryData
    data = [mData[0]]

    for mRow in mData[1:] :
        # 같은 한글이 있는지 확인
        row_data = [mRow[0]]
        for eRow in eData :
            if len(mRow) > 1 and mRow[-1] != '':
                row_data.extend([mRow[1], '√'])
                break

            elif mRow[0].replace(' ', '') == eRow[0].replace(' ', '') and eRow[1] != '' :
                row_data.extend([eRow[1], '√'])
                break

        data.append(row_data)

    print("비교 완료")
    return data


def Efind(eGlossaryData, miniGlossaryData) :
    print("엑셀 파일에서 용어가 포함된 내용을 검색 중...")
    eData = eGlossaryData
    mData = miniGlossaryData
    data = [mData[0]]

    for mRow in mData[1:] :
        # 용어가 포함된 한글이 있는지 확인
        row_data = [mRow[0]]
        for eRow in eData :
            if len(mRow) > 1 and mRow[-1] != '':
                row_data.extend([mRow[1], '√'])
                break

            elif mRow[0].replace(' ', '') in eRow[0].replace(' ', '') and eRow[1] != '' :
                row_data.append(eRow[1])
                break

        data.append(row_data)

    print("검색 완료")
    return data


def connectGsheet(googleGlossaryPath) :
    print("구글 시트에 접속 중..")
    dirPath = str(pathlib.Path.cwd())

    # json_file_name = dirPath + '/majestic-layout-275109-d180b5dbabbe.json'
    json_file_name = dirPath + '\\majestic-layout-275109-d180b5dbabbe.json'
    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive',
    ]
    try :
        credentials = ServiceAccountCredentials.from_json_keyfile_name(json_file_name, scope)
        gc = gspread.authorize(credentials)
        doc = gc.open_by_url(googleGlossaryPath)
    except Exception as e :
        print(e)
        print("[Error] 권한이 없습니다.")
        os.system("pause")

    return doc


def loadMGlossary(path, setting) :
    data = []
    # 미니 용어집 열 정보 받아오기
    selected_mrow = setting[1]
    if '.csv' not in path :
        print('[Error] 잘못된 파일입니다.')
        path = ''
    else :
        path = path.replace('"','').replace('& ','').replace("'",'')
        
        print("미니 용어집 확인 중..")
        data = []
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8-sig') as csvFile:
                csvReader = list(csv.reader(csvFile))

        # 한글, 타겟 언어만 가져오기
        # 22.10.21 index 오류로 인한 수정 - 예외처리
        try:
            for row in csvReader :
                data.append([row[int(selected_mrow[0])], row[int(selected_mrow[1])]])
        except IndexError as e:
            print('[Error] 미니용어집의 한글, 타겟열 설정이 잘못되었습니다.\n프로그램을 종료합니다.')
            os.system("pause")
            exit()

        print("미니 용어집 확인 완료")
        return data


def load_setting() :
    dirPath = str(pathlib.Path.cwd())
    setting = []
    print("setting 파일 확인 중..")
    if os.path.exists(f'{dirPath}/glosarry_downloader_settings.csv'):
        with open(f'{dirPath}/glosarry_downloader_settings.csv', 'r', encoding='utf-8-sig') as csvFile:
            csvReader = list(csv.reader(csvFile))
            # 제대로 된 파일이면 setting 데이터를 반환
            if all(
                [bool('미니 용어집 열 정보' in csvReader[0]),
                bool('텀베이스 용어집 주소' in csvReader[3]),
                bool('텀베이스 용어집 열 정보' in csvReader[6]),
                bool('엑셀 파일 열 정보' in csvReader[9]),
                bool(len(csvReader[1]) == 2),
                bool((len(csvReader[4]) == 1 and len(csvReader[7]) == 2) or 
                    (len(csvReader[4]) == 0 and len(csvReader[7]) == 2) or 
                    (len(csvReader[4]) == 0 and len(csvReader[7]) == 0)
                    ),
                bool(len(csvReader[10]) == 2 or len(csvReader[10]) == 4),]
                ) :
                print("setting 파일을 확인했습니다.")
                for row in csvReader :
                    setting.append(row)
            else :
                print("\n***** [Notice] 잘못된 데이터 구성입니다.\n설정을 진행합니다.")
                setting = setSetting()
                writeFile(setting, 'glosarry_downloader_settings')
                return setting
    else :
        print("\n***** [Notice] 설정 파일이 없습니다. *****\n설정을 진행합니다.")
        setting = setSetting()
        writeFile(setting, 'glosarry_downloader_settings')

    return setting
    pass


def selectRow() :
    source_row = input('소스 열 (a~z) : ')
    source_row = ord(source_row.upper().strip()) - 65

    taget_row = input('타겟 열 (a~z) : ')
    taget_row = ord(taget_row.upper().strip()) - 65
    print('')
    return [source_row, taget_row]


def setSetting() :
    setting = []
    print("안내에 따라 정보를 입력하세요.")

    # 미니 용어집 열 정보 설정
    setting.append(['미니 용어집 열 정보'])
    print('미니 용어집 열 설정')
    selcted_mrow = selectRow()
    setting.append(selcted_mrow)
    setting.append('')

    # 텀베이스 용어집 주소 설정
    setting.append(['텀베이스 용어집 주소'])
    setting.append([input('텀베이스 용어집 주소 입력\n> ')])
    if setting[4] == ['']:
        setting.pop()
        setting.append('')
    setting.append('')
    print('')

    # 텀베이스 용어집 열 정보 설정
    setting.append(['텀베이스 용어집 열 정보'])
    if setting[4]:
        print('텀베이스 용어집 열 설정')
        selcted_grow = selectRow()
        setting.append(selcted_grow)
    else:
        print('텀베이스 용어집 주소가 없어 열 설정을 생략합니다.')
        setting.append('')
    setting.append('')

    # 엑셀 파일 열 정보 설정
    setting.append(['엑셀 파일 열 정보'])
    print('엑셀 파일 열 설정')
    selcted_grow = selectRow()
    setting.append(selcted_grow)
    setting.append('')

    # 추가적인 엑셀 파일 열 정보 설정
    setting.append(['예외 엑셀 파일 열 정보'])
    while True:
        answer = input('예외 엑셀 파일에 대한 열 설정을 추가로 진행 하시겠습니까?\ny 또는 n을 입력하세요.\n')
        if answer == 'n':
            setting.append('')
            break
        else:
            tmp = [input('예외 엑셀 파일에 *공통적으로 포함되는 파일명* 입력\n')]
            selcted_grow = selectRow()
            tmp.extend(selcted_grow)
            setting.append(tmp)


    print("설정 완료")
    return setting


def writeFile(data, filename, dirpath = '') :
    if dirpath == '':
        dirPath = str(pathlib.Path.cwd()) + f'/{filename}.csv'
    else:
        dirPath = str(dirpath) + f'/{filename}.csv'

    print(f"{filename}.csv 파일을 내보내는 중..")
    with open(dirPath, 'w', encoding='utf-8-sig', newline='') as writeFile:
        try:
            csvWriter = csv.writer(writeFile)
            csvWriter.writerows(data)
        except Exception as e:
            print(e)

if __name__ == '__main__' :
    run()