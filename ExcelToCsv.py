import csv
import os
from os import walk
from openpyxl import load_workbook

def run(fileAbsolutePath) :
    changedPaths = []
    if os.path.splitext(fileAbsolutePath)[1].lower() == ".xlsx":
        loadedFile = load_workbook(fileAbsolutePath, data_only=True)
        sheetNames = loadedFile.sheetnames
        for sheetName in sheetNames:
            data = []
            sheet = loadedFile[sheetName]
            for row in sheet.iter_rows(min_row=1):
                row_value = []
                for cell in row:
                    try :
                        if chr(10) in cell.value :
                            cell.value = cell.value.replace(chr(10), '\t')
                    except TypeError :
                        pass
                    row_value.append(cell.value)
                data.append(row_value)

            # DirName = fileAbsolutePath.split('\\')[-2]
            fileName = fileAbsolutePath.split('\\')[-1]
            # changedPath = fileAbsolutePath.replace(DirName, DirName + "\\" + DirName + "_csv", 1)
            # if not os.path.exists(os.path.dirname(changedPath)):
            #     os.makedirs(os.path.dirname(changedPath))
            changedPath = fileAbsolutePath.replace(fileName, fileName +"_" + sheetName + ".csv")
            a =1


            with open(changedPath, 'w', encoding="utf-8-sig", newline='') as writeFile:
                try:
                    csvWriter = csv.writer(writeFile)
                    csvWriter.writerows(data)
                except Exception as e:
                    print(e)

            changedPaths.append(changedPath)
    return changedPaths

if __name__ == '__main__' :
    DirPath = input('파일 경로 입력\n')
    DirPath = DirPath.replace('"', '')
    # DirPath = 'D:\\!Project\\Umoo\\umoo\\KingsRaid\\Work\\0628_0706 에르제 소재 전면팝업 이미지 번역_7개언어\\test'
    # List 안에 파일 이름들을 모두 추가해주는 코드
    targetFilesAbsolutePaths = []
    for (dirPaths, dirNames, fileNames) in walk(DirPath):
        targetFilesAbsolutePaths.extend([dirPaths + '\\' + fileName for fileName in fileNames])

    for fileAbsolutePath in targetFilesAbsolutePaths:
            run(fileAbsolutePath)